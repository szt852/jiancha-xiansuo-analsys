from fastapi import FastAPI, File, UploadFile, Request, HTTPException
from fastapi.templating import Jinja2Templates
from fastapi.responses import StreamingResponse, HTMLResponse, Response
from fastapi.staticfiles import StaticFiles
import pandas as pd
import io
import os
from datetime import datetime
import re
from starlette.middleware.base import BaseHTTPMiddleware

# 创建FastAPI应用
app = FastAPI(title="劳动监察线索表格汇聚和统计工具")

# 添加中间件处理Vite客户端请求
class ViteClientMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        if request.url.path == "/@vite/client":
            # 静默处理Vite客户端请求，返回空响应
            return Response(content="", status_code=204)
        
        # 正常处理其他请求
        response = await call_next(request)
        return response

app.add_middleware(ViteClientMiddleware)

# 配置静态文件目录和模板
templates = Jinja2Templates(directory="templates")
app.mount("/static", StaticFiles(directory="static"), name="static")

# 区域转换映射
district_mapping = {
    "长阳土家族自治县": "长阳县",
    "五峰土家族自治县": "五峰县",
    "宜都": "宜都市",
    "当阳": "当阳市",
    "枝江": "枝江市",
    "点军": "点军区",
    "高新": "高新区",
    "西陵": "西陵区",
    "猇亭": "猇亭区",
    "兴山": "兴山县",
    "长阳": "长阳县",
    "秭归": "秭归县",
    "五峰": "五峰县",
    # 可以根据需要添加更多映射
}

# 有效区域列表
valid_districts = ["宜都市", "枝江市", "当阳市", "远安县", "兴山县", "秭归县", "长阳县", "五峰县", "夷陵区", "西陵区", "伍家岗区", "点军区", "猇亭区", "高新区"]

# 处理安薪在线数据中的区域格式
def process_anxin_district(district_str):
    """从'宜昌市-伍家岗区-伍家乡'或'宜昌市-宜都市'格式提取区域名"""
    # 分割字符串并获取倒数第二部分（如果格式为'宜昌市-伍家岗区-伍家乡'）
    # 或者获取最后一部分（如果格式为'宜昌市-宜都市'）
    parts = district_str.split('-')
    if len(parts) == 3:
        # 格式为'宜昌市-伍家岗区-伍家乡'，取中间部分
        district = parts[1]
    elif len(parts) == 2:
        # 格式为'宜昌市-宜都市'，取最后部分
        district = parts[1]
    else:
        district = district_str
    
    # 应用映射转换
    if district in district_mapping:
        return district_mapping[district]
    
    # 确保返回的是有效区域，如果不是，尝试提取有效部分
    for valid_district in valid_districts:
        if valid_district in district:
            return valid_district
    
    return district

# 处理12345数据中的区域格式
def process_12345_district(district_str):
    """标准化12345数据中的区域名称"""
    # 应用映射转换
    for key, value in district_mapping.items():
        if key in district_str:
            return value
    
    # 确保返回的是有效区域，如果不是，尝试提取有效部分
    for valid_district in valid_districts:
        if valid_district in district_str:
            return valid_district
    
    return district_str

# 处理12345数据
def process_12345_data(df):
    """处理12345数据并转换为汇总表格式"""
    # 确保必要的列存在
    required_columns = ["序号", "流水号", "紧急程度", "事件来源", "诉求人", "诉求人电话", 
                       "诉求标题", "诉求内容", "所属区域", "所涉领域", "是否涉稳", 
                       "所涉行业", "所涉项目（企业）", "项目性质", "是否在监管系统中", 
                       "项目状态", "建设单位", "总包单位", "涉及人数", "涉及金额", "是否再次投诉"]
    
    for col in required_columns:
        if col not in df.columns:
            # 如果列不存在，添加空列
            df[col] = pd.NA
    
    # 标准化所属区域
    df['所属区域'] = df['所属区域'].apply(process_12345_district)
    
    # 对于12345数据，需要添加安薪在线数据相关的空列
    anxin_columns = ["施工单位", "项目经理", "联系电话", "预警类型", "预警原因", 
                    "预警时间", "状态", "预警天数"]
    
    for col in anxin_columns:
        df[col] = pd.NA
    
    return df

# 处理安薪在线数据
def process_anxin_data(df):
    """处理安薪在线数据并转换为汇总表格式"""
    # 确保必要的列存在
    required_columns = ["项目名称", "行业", "建设单位", "施工单位", "项目经理", 
                       "联系电话", "区域", "预警类型", "预警原因", "预警时间", 
                       "状态", "预警天数"]
    
    for col in required_columns:
        if col not in df.columns:
            # 如果列不存在，添加空列
            df[col] = pd.NA
    
    # 创建新的数据框，用于汇总表格式
    result_df = pd.DataFrame()
    
    # 填充汇总表所需的字段
    result_df['序号'] = range(1, len(df) + 1)
    result_df['流水号'] = pd.NA
    result_df['紧急程度'] = pd.NA
    result_df['事件来源'] = "安薪在线"  # 安薪在线数据的事件来源固定为"安薪在线"
    result_df['诉求人'] = pd.NA
    result_df['诉求人电话'] = pd.NA
    result_df['诉求标题'] = pd.NA
    result_df['诉求内容'] = pd.NA
    result_df['所属区域'] = df['区域'].apply(process_anxin_district)  # 转换区域格式
    result_df['所涉领域'] = "建筑"  # 安薪在线数据的所涉领域固定为"建筑"
    result_df['是否涉稳'] = pd.NA
    result_df['所涉行业'] = df['行业']
    result_df['所涉项目（企业）'] = df['项目名称']
    result_df['项目性质'] = pd.NA
    result_df['是否在监管系统中'] = pd.NA
    result_df['项目状态'] = pd.NA
    result_df['建设单位'] = df['建设单位']
    result_df['总包单位'] = pd.NA
    result_df['涉及人数'] = pd.NA
    result_df['涉及金额'] = pd.NA
    result_df['是否再次投诉'] = pd.NA
    result_df['施工单位'] = df['施工单位']
    result_df['项目经理'] = df['项目经理']
    result_df['联系电话'] = df['联系电话']
    result_df['预警类型'] = df['预警类型']
    result_df['预警原因'] = df['预警原因']
    result_df['预警时间'] = df['预警时间']
    result_df['状态'] = df['状态']
    result_df['预警天数'] = df['预警天数']
    
    return result_df

# 生成汇总表
def generate_summary_table(df_12345, df_anxin):
    """合并12345数据和安薪在线数据生成汇总表"""
    # 处理两个数据源
    processed_df_12345 = process_12345_data(df_12345)
    processed_df_anxin = process_anxin_data(df_anxin)
    
    # 合并两个数据框
    combined_df = pd.concat([processed_df_12345, processed_df_anxin], ignore_index=True)
    
    # 重新生成序号
    combined_df['序号'] = range(1, len(combined_df) + 1)
    
    return combined_df

# 安全地将值转换为整数的辅助函数
def safe_int_convert(value):
    try:
        # 检查是否为字符串，如果是则去除空格
        if isinstance(value, str):
            value = value.strip()
            # 如果是空格或空字符串，返回0
            if not value:
                return 0
        # 尝试转换为整数
        return int(value)
    except (ValueError, TypeError):
        # 转换失败返回0
        return 0

# 生成基本情况文本
def generate_basic_info(summary_df, current_date):
    """根据汇总表生成基本情况文本"""
    # 准备数据
    # 1. 线索数量相关统计
    non_anxin_data = summary_df[summary_df['事件来源'] != '安薪在线']
    anxin_data = summary_df[summary_df['事件来源'] == '安薪在线']
    
    # 1.1 日期格式化为YYYY年MM月DD日
    datetime_str = f"{current_date[:4]}年{current_date[4:6]}月{current_date[6:]}日"
    
    # 1.2 总线索数量
    xiansuo_count = len(non_anxin_data)
    
    # 1.3 统计所有事件来源的数量
    event_source_counts = non_anxin_data['事件来源'].value_counts()
    
    # 1.4 部平台线索数量
    xiansuo_bupingtai_count = len(non_anxin_data[non_anxin_data['事件来源'] == '部平台'])
    
    # 1.5 12345热线线索数量
    xiansuo_12345_count = len(non_anxin_data[non_anxin_data['事件来源'].str.startswith('12345/')])
    
    # 1.5 一人多诉统计
    phone_counts = non_anxin_data['诉求人电话'].value_counts()
    yirenduosu_count = len(phone_counts[phone_counts >= 3])
    yirenduosu = f"{yirenduosu_count}条" if yirenduosu_count > 0 else "无"
    
    # 1.6 多人一诉统计
    project_counts = non_anxin_data['所涉项目（企业）'].value_counts()
    duorenyisu_count = len(project_counts[project_counts >= 3])
    duorenyisu = f"{duorenyisu_count}条" if duorenyisu_count > 0 else "无"
    
    # 1.7 再次投诉统计
    # 查找“是否再次投诉”字段值为“是”的数据
    zaicitousu_count = len(non_anxin_data[non_anxin_data['是否再次投诉'].str.contains('是', na=False)])
    
    # 1.8 线索数量前三的地区
    full_district_counts = non_anxin_data['所属区域'].value_counts()
    district_counts_top3 = full_district_counts.head(3)
    xiansuo_count_top3 = "、".join([f"{district}{count}条" for district, count in district_counts_top3.items()])
    
    # 第一段文字
    text1 = f"1.线索数量。{datetime_str}，全市共收到欠薪线索{xiansuo_count}条（部平台{xiansuo_bupingtai_count}条、12345热线{xiansuo_12345_count}条），较上一期减少[手工填写]条（部平台减少[手工填写]条）。一人多诉{yirenduosu}，多人一诉{duorenyisu}，再次投诉{zaicitousu_count}条。线索数量前三的地方为：{xiansuo_count_top3}。"
    
    # 2. 建设领域欠薪线索情况
    construction_data = non_anxin_data[non_anxin_data['所涉领域'] == '建筑']
    jianshe_project_count = len(construction_data)
    
    # 2.2 建设领域行业分布
    industry_counts = construction_data['所涉行业'].value_counts()
    jianshe_suoshe_hangye = "其中" + "、".join([f"{industry}{count}条" for industry, count in industry_counts.items()])
    
    # 2.3 建设领域项目性质统计
    government_count = len(construction_data[construction_data['项目性质'].str.contains('政府', na=False)])
    state_owned_count = len(construction_data[construction_data['项目性质'].str.contains('国企', na=False)])
    other_count = len(construction_data[~construction_data['项目性质'].str.contains('政府|国企', na=False)])
    jianshe_project_xingzhi = f"涉及政府项目{government_count}个、国企项目{state_owned_count}个、其他商建项目{other_count}个"
    
    # 2.4 建设领域涉及人数
    jianshe_renshu = construction_data['涉及人数'].apply(safe_int_convert).sum()
    
    # 2.5 建设领域涉及金额（万元）
    jianshe_jine = round(construction_data['涉及金额'].apply(safe_int_convert).sum() / 10000, 2)
    
    # 第二段文字
    text2 = f"2.建设领域欠薪线索情况。建设领域{jianshe_project_count}条。({jianshe_suoshe_hangye}，{jianshe_project_xingzhi}），涉及{jianshe_renshu}人、{jianshe_jine}万元。"
    
    # 3. 非建领域欠薪线索情况
    non_construction_data = non_anxin_data[non_anxin_data['所涉领域'] == '非建']
    feijian_project_count = len(non_construction_data)
    
    # 3.2 非建领域行业分布
    feijian_industry_counts = non_construction_data['所涉行业'].value_counts()
    feijian_suoshe_hangye = "其中" + "、".join([f"{industry}{count}条" for industry, count in feijian_industry_counts.items()])
    
    # 转换为前端可用的格式
    feijian_industry_data = {industry: count for industry, count in feijian_industry_counts.items()}
    
    # 3.3 非建领域涉及人数
    feijian_renshu = non_construction_data['涉及人数'].apply(safe_int_convert).sum()
    
    # 3.4 非建领域涉及金额（万元）
    feijian_jine = round(non_construction_data['涉及金额'].apply(safe_int_convert).sum() / 10000, 2)
    
    # 第三段文字
    text3 = f"3.非建领域欠薪线索情况。非建领域{feijian_project_count}条。（{feijian_suoshe_hangye}），涉及{feijian_renshu}人、{feijian_jine}万元。"
    
    # 4. 涉稳情况（固定格式）
    text4 = "4.涉稳情况。舆情[手工填写]条。[手工填写]"
    
    # 5. 预警情况
    axzx_yvjing_count = len(anxin_data)
    
    # 5.2 预警类型统计
    warning_types = anxin_data['预警类型'].value_counts()
    axzx_yvjing_leixing = "其中" + "；".join([f"{warning}{count}条" for warning, count in warning_types.items()])
    
    # 第五段文字
    text5 = f"5.预警情况。安薪在线系统产生预警信息{axzx_yvjing_count}条，{axzx_yvjing_leixing}。"
    
    # 合并所有文字，用换行符分隔
    basic_info = "\n".join([text1, text2, text3, text4, text5])
    
    # 按县市区统计建设领域线索数量
    jianshe_district_counts = construction_data['所属区域'].value_counts()
    
    # 按县市区统计非建领域线索数量
    feijian_district_counts = non_construction_data['所属区域'].value_counts()
    
    # 按县市区统计预警信息数量
    warning_district_counts = anxin_data['所属区域'].value_counts()
    
    # 建筑类其他维度分析数据
    # 1. 事件来源统计（排除安薪在线）
    jianshe_event_source_counts = construction_data['事件来源'].value_counts()
    
    # 2. 所涉行业统计（排除安薪在线）
    jianshe_industry_counts = construction_data['所涉行业'].value_counts()
    
    # 3. 项目性质统计
    jianshe_project_nature_counts = construction_data['项目性质'].value_counts()
    
    # 4. 涉及人数金额数据（用于散点图）
    jianshe_scatter_data = []
    for idx, row in construction_data.iterrows():
        people = safe_int_convert(row['涉及人数'])
        amount = safe_int_convert(row['涉及金额'])
        if people > 0 or amount > 0:
            jianshe_scatter_data.append({
                'id': idx + 1,
                'people': people,
                'amount': amount
            })
    
    # 计算人数和金额的平均值
    jianshe_people_avg = sum(item['people'] for item in jianshe_scatter_data) / len(jianshe_scatter_data) if jianshe_scatter_data else 0
    jianshe_amount_avg = sum(item['amount'] for item in jianshe_scatter_data) / len(jianshe_scatter_data) if jianshe_scatter_data else 0
    
    # 非建类其他维度分析数据
    # 1. 事件来源统计
    feijian_event_source_counts = non_construction_data['事件来源'].value_counts()
    
    # 2. 所涉行业统计（已存在，使用feijian_industry_counts）
    
    # 3. 涉及人数金额数据（用于散点图）
    feijian_scatter_data = []
    for idx, row in non_construction_data.iterrows():
        people = safe_int_convert(row['涉及人数'])
        amount = safe_int_convert(row['涉及金额'])
        if people > 0 or amount > 0:
            feijian_scatter_data.append({
                'id': idx + 1,
                'people': people,
                'amount': amount
            })
    
    # 计算人数和金额的平均值
    feijian_people_avg = sum(item['people'] for item in feijian_scatter_data) / len(feijian_scatter_data) if feijian_scatter_data else 0
    feijian_amount_avg = sum(item['amount'] for item in feijian_scatter_data) / len(feijian_scatter_data) if feijian_scatter_data else 0
    
    # 建筑类涉及人数较多的项目数据（涉及人数>3）
    jianshe_large_projects = construction_data[construction_data['涉及人数'].apply(safe_int_convert) >= 3].copy()
    jianshe_large_projects_list = []
    
    # 提取所需字段并处理空值
    for _, row in jianshe_large_projects.iterrows():
        project_data = {
            'project_name': str(row['所涉项目（企业）']) if pd.notna(row['所涉项目（企业）']) and str(row['所涉项目（企业）']).strip() else '--',
            'district': str(row['所属区域']) if pd.notna(row['所属区域']) and str(row['所属区域']).strip() else '--',
            'industry': str(row['所涉行业']) if pd.notna(row['所涉行业']) and str(row['所涉行业']).strip() else '--',
            'people_count': int(safe_int_convert(row['涉及人数'])),
            'amount': float(safe_int_convert(row['涉及金额'])),
            'applicant': str(row['诉求人']) if pd.notna(row['诉求人']) and str(row['诉求人']).strip() else '--',
            'content': str(row['诉求内容']) if pd.notna(row['诉求内容']) and str(row['诉求内容']).strip() else '--'
        }
        jianshe_large_projects_list.append(project_data)
    
    # 非建类涉及人数较多的项目数据（涉及人数>3）
    feijian_large_projects = non_construction_data[non_construction_data['涉及人数'].apply(safe_int_convert) > 3].copy()
    feijian_large_projects_list = []
    
    # 提取所需字段并处理空值
    for _, row in feijian_large_projects.iterrows():
        project_data = {
            'project_name': str(row['所涉项目（企业）']) if pd.notna(row['所涉项目（企业）']) and str(row['所涉项目（企业）']).strip() else '--',
            'district': str(row['所属区域']) if pd.notna(row['所属区域']) and str(row['所属区域']).strip() else '--',
            'industry': str(row['所涉行业']) if pd.notna(row['所涉行业']) and str(row['所涉行业']).strip() else '--',
            'people_count': int(safe_int_convert(row['涉及人数'])),
            'amount': float(safe_int_convert(row['涉及金额'])),
            'applicant': str(row['诉求人']) if pd.notna(row['诉求人']) and str(row['诉求人']).strip() else '--',
            'content': str(row['诉求内容']) if pd.notna(row['诉求内容']) and str(row['诉求内容']).strip() else '--'
        }
        feijian_large_projects_list.append(project_data)
    
    # 准备所有县市区列表，确保所有分类都有完整的县市区数据
    all_districts = set(full_district_counts.index)
    
    # 预警类统计数据
    # 1. 预警类案件总数（事件来源为"安薪在线"的数据数量）
    warning_case_total = len(anxin_data)
    
    # 2. 涉及建设单位（统计建设单位字段中出现的唯一值数量）
    unique_construction_units = anxin_data['建设单位'].dropna().nunique()
    
    # 3. 所涉项目/企业（统计所涉项目（企业）字段中出现的唯一值数量）
    unique_projects = anxin_data['所涉项目（企业）'].dropna().nunique()
    
    # 4. 所涉行业统计
    warning_industry_counts = anxin_data['所涉行业'].value_counts()
    
    # 5. 预警类型统计（已存在）
    # warning_types 已定义
    
    # 6. 状态统计
    warning_status_counts = anxin_data['状态'].value_counts()
    
    # 准备数据看板数据，确保所有数值都是Python原生类型
    dashboard_data = {
        'datetime': datetime_str,
        'basic_info': basic_info,  # 添加基本情况文本
        'xiansuo_count': int(xiansuo_count),
        'xiansuo_bupingtai_count': int(xiansuo_bupingtai_count),
        'xiansuo_12345_count': int(xiansuo_12345_count),
        'yirenduosu': yirenduosu,
        'duorenyisu': duorenyisu,
        'zaicitousu_count': int(zaicitousu_count),
        'xiansuo_count_top3': xiansuo_count_top3,
        'jianshe_project_count': int(jianshe_project_count),
        'jianshe_renshu': int(jianshe_renshu),
        'jianshe_jine': float(jianshe_jine),
        'feijian_project_count': int(feijian_project_count),
        'feijian_renshu': int(feijian_renshu),
        'feijian_jine': float(feijian_jine),
        'feijian_industry_data': feijian_industry_data,
        'axzx_yvjing_count': int(axzx_yvjing_count),
        # 预警类统计数据
        'warning_case_total': int(warning_case_total),
        'unique_construction_units': int(unique_construction_units),
        'unique_projects': int(unique_projects),
        'warning_industry_counts': {k: int(v) for k, v in warning_industry_counts.to_dict().items()},
        'warning_status_counts': {k: int(v) for k, v in warning_status_counts.to_dict().items()},
        # 将Series.to_dict()的结果转换为Python原生类型的字典
        'district_counts': {k: int(v) for k, v in full_district_counts.to_dict().items()},
        'jianshe_district_counts': {k: int(jianshe_district_counts[k]) if k in jianshe_district_counts else 0 for k in all_districts},
        'feijian_district_counts': {k: int(feijian_district_counts[k]) if k in feijian_district_counts else 0 for k in all_districts},
        'warning_district_counts': {k: int(warning_district_counts[k]) if k in warning_district_counts else 0 for k in all_districts},
        'industry_counts': {k: int(v) for k, v in industry_counts.to_dict().items()},
        'feijian_industry_counts': {k: int(v) for k, v in feijian_industry_counts.to_dict().items()},
        'warning_types': {k: int(v) for k, v in warning_types.to_dict().items()},
        'project_nature_counts': {
            '政府项目': int(government_count),
            '国企项目': int(state_owned_count),
            '其他商建项目': int(other_count)
        },
        'event_source_counts': {k: int(v) for k, v in event_source_counts.to_dict().items()},
        # 建筑类其他维度分析数据
        'jianshe_event_source_counts': {k: int(v) for k, v in jianshe_event_source_counts.to_dict().items()},
        'jianshe_industry_counts': {k: int(v) for k, v in jianshe_industry_counts.to_dict().items()},
        'jianshe_project_nature_counts': {k: int(v) for k, v in jianshe_project_nature_counts.to_dict().items()},
        'jianshe_scatter_data': jianshe_scatter_data,
        'jianshe_people_avg': float(jianshe_people_avg),
        'jianshe_amount_avg': float(jianshe_amount_avg),
        # 非建类其他维度分析数据
        'feijian_event_source_counts': {k: int(v) for k, v in feijian_event_source_counts.to_dict().items()},
        'feijian_scatter_data': feijian_scatter_data,
        'feijian_people_avg': float(feijian_people_avg),
        'feijian_amount_avg': float(feijian_amount_avg),
        # 涉及人数较多的项目数据
        'jianshe_large_projects': jianshe_large_projects_list,
        'feijian_large_projects': feijian_large_projects_list
    }
    
    return basic_info, dashboard_data

# 生成数据情况统计表
def generate_statistics_table(summary_df, basic_info):
    """根据汇总表生成数据情况统计表"""
    # 创建统计结果数据框
    stats_data = []
    
    # 添加基本情况行
    # stats_data.append({
    #     '基本情况': basic_info,
    #     '区域': '',
    #     '专班包保（后续有调整）': '',
    #     '领域': '',
    #     '线索数量': '',
    #     '涉稳': '',
    #     '预警': '',
    #     '欠薪点位': ''
    # })
    
    # 区域到专班包保人的映射关系
    district_boss_mapping = {
        "宜都市": "罗雷",
        "枝江市": "孟禹",
        "当阳市": "侯民杰",
        "远安县": "熊伟",
        "兴山县": "牟鹏",
        "秭归县": "叶磊",
        "长阳县": "杨继平",
        "五峰县": "肖丰",
        "夷陵区": "韩晓明",
        "西陵区": "董蒋军",
        "伍家岗区": "雷斌斌",
        "点军区": "储成刚",
        "猇亭区": "朱强",
        "高新区": "侯民杰"
    }
    
    row_number=0
    # 对于每个有效区域，创建建设和非建两行数据
    for district in valid_districts:
        if row_number==0:
            basic_info=basic_info
        else:
            basic_info=""
        row_number=row_number+1

        # 建设领域数据统计 - 只统计预警类型字段文本长度为0的记录
        construction_data = summary_df[(summary_df['所属区域'] == district) & 
                                      (summary_df['所涉领域'] == '建筑')]
        
        # 非建领域数据统计 - 只统计预警类型字段文本长度为0的记录
        non_construction_data = summary_df[(summary_df['所属区域'] == district) & 
                                          (summary_df['所涉领域'] != '建筑')]
        
        # 统计建设领域的线索数量 - 只统计预警类型字段文本长度为0的记录
        # 处理预警类型字段，确保是字符串类型并检查长度
        construction_with_empty_warning = construction_data[
            construction_data['预警类型'].apply(lambda x: isinstance(x, str) and len(x.strip()) == 0 or pd.isna(x))
        ]
        construction_count = len(construction_with_empty_warning)
        
        # 统计非建领域的线索数量 - 只统计预警类型字段文本长度为0的记录
        non_construction_with_empty_warning = non_construction_data[
            non_construction_data['预警类型'].apply(lambda x: isinstance(x, str) and len(x.strip()) == 0 or pd.isna(x))
        ]
        non_construction_count = len(non_construction_with_empty_warning)
        
        # 统计预警数量（只有建设领域有预警）
        warning_count = len(construction_data[construction_data['预警类型'].notna()])
        
        # 获取当前区域对应的专班包保人
        boss_name = district_boss_mapping.get(district, "")
        
        # 已将safe_int_convert移至全局函数
        
        # 构建建设领域的欠薪点位文本
        # 1. 处理无预警的建设领域数据
        construction_no_warning = construction_data[construction_data['预警类型'].isna()]
        wage_arrears_construction_no_warning = ""
        for _, row in construction_no_warning.iterrows():
            project = str(row['所涉项目（企业）']) if pd.notna(row['所涉项目（企业）']) else ""
            people = str(safe_int_convert(row['涉及人数'])) if pd.notna(row['涉及人数']) else "0"
            amount = str(safe_int_convert(row['涉及金额'])) if pd.notna(row['涉及金额']) else "0"
            wage_arrears_construction_no_warning += f"{project}（{people}人{amount}元）；"
        
        # 2. 处理有预警的建设领域数据
        construction_with_warning = construction_data[construction_data['预警类型'].notna()]
        wage_arrears_construction_with_warning = ""
        for _, row in construction_with_warning.iterrows():
            project = str(row['所涉项目（企业）']) if pd.notna(row['所涉项目（企业）']) else ""
            reason = str(row['预警原因']) if pd.notna(row['预警原因']) else ""
            wage_arrears_construction_with_warning += f"{project}因{reason}预警；"
        
        # 合并建设领域的欠薪点位
        wage_arrears_construction = wage_arrears_construction_no_warning + wage_arrears_construction_with_warning
        
        # 构建非建领域的欠薪点位文本（非建领域没有预警）
        wage_arrears_non_construction = ""
        for _, row in non_construction_data.iterrows():
            project = str(row['所涉项目（企业）']) if pd.notna(row['所涉项目（企业）']) else ""
            people = str(safe_int_convert(row['涉及人数'])) if pd.notna(row['涉及人数']) else "0"
            amount = str(safe_int_convert(row['涉及金额'])) if pd.notna(row['涉及金额']) else "0"
            wage_arrears_non_construction += f"{project}（{people}人{amount}元）；"
        
        # 添加建设领域的数据行
        stats_data.append({
            '基本情况': basic_info,
            '区域': district,
            '专班包保（后续有调整）': boss_name,  # 根据区域自动填充专班包保人
            '领域': '建设',
            '线索数量': construction_count,
            '涉稳': '无',
            '预警': warning_count,
            '欠薪点位': wage_arrears_construction
        })
        
        # 添加非建领域的数据行
        stats_data.append({
            '基本情况': '',
            '区域': '',
            '专班包保（后续有调整）': boss_name,  # 非建领域也使用相同的专班包保人
            '领域': '非建',
            '线索数量': non_construction_count,
            '涉稳': '无',
            '预警': 0,
            '欠薪点位': wage_arrears_non_construction
        })
    
    # 创建统计结果数据框
    stats_df = pd.DataFrame(stats_data)
    
    return stats_df

@app.get("/")
async def read_root(request: Request):
    """返回主页面"""
    return templates.TemplateResponse("index.html", {"request": request})

# @app.get("/test")
# async def test_page(request: Request):
#     return templates.TemplateResponse("test.html", {"request": request})

# @app.post("/test_upload/")
# async def test_upload(file_12345: UploadFile = File(...), file_anxin: UploadFile = File(...)):
#     """测试文件上传功能的端点"""
#     print(f"测试上传端点收到请求：{file_12345.filename}, {file_anxin.filename}")
#     try:
#         # 简单读取文件内容以验证上传
#         file_12345_content = await file_12345.read()
#         file_anxin_content = await file_anxin.read()
#         print(f"文件内容读取成功，大小分别为：{len(file_12345_content)} 字节, {len(file_anxin_content)} 字节")
#         return {"status": "success", "message": "文件上传成功", "files": [file_12345.filename, file_anxin.filename]}
#     except Exception as e:
#         print(f"测试上传时出错: {str(e)}")
#         import traceback
#         print(traceback.format_exc())
#         raise HTTPException(status_code=500, detail=f"测试上传失败: {str(e)}")

@app.post("/process_files/")
async def process_files(file_12345: UploadFile = File(...), file_anxin: UploadFile = File(...)):
    print("收到文件上传请求")
    
    # 检查文件类型
    if not file_12345.filename.endswith(('.xlsx', '.xls')):
        print(f"文件类型错误: {file_12345.filename}")
        raise HTTPException(status_code=400, detail=f"请上传Excel文件(.xlsx或.xls格式)，当前文件: {file_12345.filename}")
    
    if not file_anxin.filename.endswith(('.xlsx', '.xls')):
        print(f"文件类型错误: {file_anxin.filename}")
        raise HTTPException(status_code=400, detail=f"请上传Excel文件(.xlsx或.xls格式)，当前文件: {file_anxin.filename}")
    
    try:
        # 读取12345文件
        print(f"开始读取文件: {file_12345.filename}")
        try:
            # 保存文件内容到临时BytesIO对象
            file_12345_content = await file_12345.read()
            file_12345_io = io.BytesIO(file_12345_content)
            df_12345 = pd.read_excel(file_12345_io)
            print(f"成功读取12345文件，行数: {len(df_12345)}, 列数: {len(df_12345.columns)}")
            print(f"12345文件列名: {list(df_12345.columns)}")
        except Exception as e:
            print(f"读取12345文件失败: {str(e)}")
            import traceback
            print(traceback.format_exc())
            raise HTTPException(status_code=500, detail=f"读取12345文件失败: {str(e)}")
        
        # 读取安薪文件
        print(f"开始读取文件: {file_anxin.filename}")
        try:
            # 保存文件内容到临时BytesIO对象
            file_anxin_content = await file_anxin.read()
            file_anxin_io = io.BytesIO(file_anxin_content)
            df_anxin = pd.read_excel(file_anxin_io)
            print(f"成功读取安薪文件，行数: {len(df_anxin)}, 列数: {len(df_anxin.columns)}")
            print(f"安薪文件列名: {list(df_anxin.columns)}")
        except Exception as e:
            print(f"读取安薪文件失败: {str(e)}")
            import traceback
            print(traceback.format_exc())
            raise HTTPException(status_code=500, detail=f"读取安薪文件失败: {str(e)}")
        
        # 生成汇总表
        print("开始生成汇总表...")
        try:
            summary_df = generate_summary_table(df_12345, df_anxin)
            print(f"汇总表生成完成，行数: {len(summary_df)}, 列数: {len(summary_df.columns)}")
        except Exception as e:
            print(f"生成汇总表失败: {str(e)}")
            import traceback
            print(traceback.format_exc())
            raise HTTPException(status_code=500, detail=f"生成汇总表失败: {str(e)}")
        
        # 生成日期
        current_date = datetime.now().strftime("%Y%m%d")
        
        # 生成基本情况文本和数据看板数据
        print("开始生成基本情况...")
        try:
            basic_info, dashboard_data = generate_basic_info(summary_df, current_date)
            print("基本情况生成完成")
        except Exception as e:
            print(f"生成基本情况失败: {str(e)}")
            import traceback
            print(traceback.format_exc())
            raise HTTPException(status_code=500, detail=f"生成基本情况失败: {str(e)}")
        
        # 生成统计表
        print("开始生成统计表...")
        try:
            stats_df = generate_statistics_table(summary_df, basic_info)
            print(f"统计表生成完成，行数: {len(stats_df)}, 列数: {len(stats_df.columns)}")
        except Exception as e:
            print(f"生成统计表失败: {str(e)}")
            import traceback
            print(traceback.format_exc())
            raise HTTPException(status_code=500, detail=f"生成统计表失败: {str(e)}")
        
        # 生成文件名
        filename = f"{current_date}劳动监察线索汇总和统计.xlsx"
        
        # 创建Excel文件
        print("开始写入Excel文件...")
        try:
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # 写入汇总表
                summary_df.to_excel(writer, sheet_name='劳动监察线索汇总表', index=False)
                
                # 写入统计表
                stats_df.to_excel(writer, sheet_name='劳动监察线索数据情况统计表', index=False)
                
                # 合并基本情况单元格（A2到A29）
                worksheet = writer.sheets['劳动监察线索数据情况统计表']
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Alignment
                
                # 基本情况列在Excel中是第一列
                basic_info_col_index = 1
                
                # 保存A2单元格的内容，用于合并后填充
                a2_value = worksheet.cell(row=2, column=basic_info_col_index).value
                
                # 合并A2到A29单元格
                worksheet.merge_cells(start_row=2, start_column=basic_info_col_index, 
                                    end_row=29, end_column=basic_info_col_index)
                
                # 用A2的内容填充合并后的单元格并设置垂直居中
                a_cell = worksheet.cell(row=2, column=basic_info_col_index)
                a_cell.value = a2_value
                a_cell.alignment = Alignment(vertical='center')
                
                # 合并B列的指定单元格对（B2~B3、B4~B5等）
                for i in range(2, 30, 2):
                    # 保存B列当前单元格的值
                    b_value = worksheet.cell(row=i, column=2).value
                    # 合并B列两个相邻单元格
                    worksheet.merge_cells(start_row=i, start_column=2, end_row=i+1, end_column=2)
                    # 用第一个单元格的内容填充并设置垂直居中
                    b_cell = worksheet.cell(row=i, column=2)
                    b_cell.value = b_value
                    b_cell.alignment = Alignment(vertical='center')
                    
                    # 保存C列当前单元格的值
                    c_value = worksheet.cell(row=i, column=3).value
                    # 合并C列两个相邻单元格
                    worksheet.merge_cells(start_row=i, start_column=3, end_row=i+1, end_column=3)
                    # 用第一个单元格的内容填充并设置垂直居中
                    c_cell = worksheet.cell(row=i, column=3)
                    c_cell.value = c_value
                    c_cell.alignment = Alignment(vertical='center')
                
                # 调整列宽以适应内容
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                    
            # 重置文件指针
            output.seek(0)
            print("Excel文件写入完成")
        except Exception as e:
            print(f"写入Excel文件失败: {str(e)}")
            import traceback
            print(traceback.format_exc())
            raise HTTPException(status_code=500, detail=f"写入Excel文件失败: {str(e)}")
        
        # 返回生成的Excel文件，需要对中文文件名进行URL编码
        from urllib.parse import quote
        import json
        encoded_filename = quote(filename)
        response = StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={encoded_filename}",
                "X-Dashboard-Data": json.dumps(dashboard_data)  # 传递数据看板数据给前端
            }
        )
        
        return response
    except HTTPException:
        # 重新抛出已经格式化的HTTPException
        raise
    except Exception as e:
        print(f"处理文件时发生未捕获的错误: {str(e)}")
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"处理文件时发生未捕获的错误: {str(e)}")

# 运行应用
# if __name__ == "__main__":
#     import uvicorn
#     uvicorn.run(app, host="0.0.0.0", port=8002)